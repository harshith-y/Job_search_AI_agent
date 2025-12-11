"""
Enhanced Job Tracker with Fancy Excel Formatting
- Google Sheets-style formatting
- Color-coded dropdowns for Application Status
- Dark green header with filters
- Alternating row colors
- Professional formatting matching user's preferred style
"""

import json
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.utils import get_column_letter


class EnhancedJobTracker:
    """Enhanced job tracker with detailed information and fancy Excel export"""

    def __init__(self, filename="job_tracker.json"):
        self.filename = filename
        self.jobs = self.load()
    
    def load(self):
        """Load jobs from JSON"""
        if os.path.exists(self.filename):
            with open(self.filename, "r") as f:
                return json.load(f)
        return {}
    
    def save(self):
        """Save jobs to JSON"""
        with open(self.filename, "w") as f:
            json.dump(self.jobs, f, indent=2)
    
    def add_job(self, job, status="new"):
        """Add a job with all details"""
        
        url = job["url"]
        
        if url in self.jobs:
            # Update existing
            self.jobs[url].update({
                "status": status,
                "last_updated": datetime.now().isoformat()
            })
        else:
            # Add new
            self.jobs[url] = {
                "title": job["title"],
                "company": job["company"],
                "location": job.get("location", "UK"),
                "city": job.get("city", job.get("location", "UK")),  # NEW
                "type": job.get("job_type", job.get("type", "Industry")),
                "url": url,
                "description": job.get("description", ""),
                "ai_summary": job.get("ai_summary", ""),
                "status": status,
                "date_found": datetime.now().strftime("%Y-%m-%d"),
                "applied_date": None,
                "response": "Not Applied",
                "notes": "",
                # NEW FIELDS:
                "requirements": job.get("requirements", []),
                "expectations": job.get("expectations", []),
                "salary": job.get("salary", "Not specified"),
                "post_date": job.get("post_date", "Recent"),
                "deadline": job.get("deadline", "Not specified"),
                "cv_required": job.get("cv_required", "Not specified"),
                "cover_letter_required": job.get("cover_letter_required", "Not specified"),
            }
        
        self.save()
    
    def update_status(self, url, status):
        """Update job status"""
        if url in self.jobs:
            self.jobs[url]["status"] = status
            self.jobs[url]["last_updated"] = datetime.now().isoformat()
            self.save()
    
    def get_jobs_by_status(self, status):
        """Get all jobs with specific status"""
        return [job for job in self.jobs.values() if job["status"] == status]
    
    def export_to_excel_fancy(self, filename="job_applications.xlsx"):
        """
        Export to Excel with Google Sheets-style formatting:
        - Dark green header with filters
        - Color-coded dropdown for Application Status
        - Rejection Reason dropdown
        - Alternating row colors
        - Professional styling
        """

        URL_COLUMN = 6  # Column F: Link to Job Req

        # Check if file exists
        if os.path.exists(filename):
            print(f"📊 Updating existing spreadsheet: {filename}")
            wb = load_workbook(filename)
            ws = wb.active

            # Get existing URLs to avoid duplicates
            existing_urls = set()
            for row in range(2, ws.max_row + 1):
                url = ws.cell(row=row, column=URL_COLUMN).value
                if url:
                    existing_urls.add(url)

            # Add only new jobs
            row_num = ws.max_row + 1
            new_count = 0

            for url, job in self.jobs.items():
                if url not in existing_urls:
                    self._add_job_row_fancy(ws, row_num, job)
                    row_num += 1
                    new_count += 1
                else:
                    # Update status if changed
                    for row in range(2, ws.max_row + 1):
                        if ws.cell(row=row, column=URL_COLUMN).value == url:
                            self._update_status_cell(ws, row, job["status"])
                            break

            print(f"   ✅ Added {new_count} new jobs")
            print(f"   ✅ Updated statuses for existing jobs")
        
        else:
            print(f"📊 Creating new spreadsheet: {filename}")
            wb = Workbook()
            ws = wb.active
            ws.title = "Job Applications"
            
            # Create header row
            self._create_fancy_header(ws)
            
            # Add all jobs
            row_num = 2
            for url, job in self.jobs.items():
                self._add_job_row_fancy(ws, row_num, job)
                row_num += 1
            
            print(f"   ✅ Added {len(self.jobs)} jobs")
        
        # Add dropdowns and formatting
        self._add_fancy_dropdowns(ws)
        self._auto_fit_columns(ws)
        
        # Save
        wb.save(filename)
        print(f"✅ Spreadsheet saved: {filename}\n")
        
        return filename
    
    def _create_fancy_header(self, ws):
        """Create Google Sheets-style header with dark green background"""

        # Columns matching the user's Google Sheets format
        headers = [
            "Company Name",        # A
            "Application Status",  # B - with dropdown
            "Role",                # C
            "Salary",              # D
            "Date Submitted",      # E
            "Link to Job Req",     # F - URL/PDF link
            "Rejection Reason",    # G - with dropdown
            "Location",            # H
            "Deadline",            # I
            "Notes",               # J
            "AI Summary",          # K
        ]

        # Dark green header (matching Google Sheets screenshot)
        header_fill = PatternFill(start_color="2E5E3E", end_color="2E5E3E", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Thin border for header
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Set row height
        ws.row_dimensions[1].height = 30

        # Enable auto-filter (creates the filter dropdown arrows)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    
    def _add_job_row_fancy(self, ws, row_num, job):
        """Add a job row matching Google Sheets format"""

        # Map internal status to display status
        status_map = {
            "new": "Have Not Applied",
            "liked": "Have Not Applied",
            "maybe": "Have Not Applied",
            "disliked": "Have Not Applied",
            "applied": "Submitted - Pending Response",
            "interview": "Submitted - Pending Response",
            "offer": "Submitted - Pending Response",
            "rejected": "Submitted - Pending Response"
        }

        application_status = status_map.get(job.get("status", "new"), "Have Not Applied")

        # If actually applied, use the applied date, otherwise use date found
        date_submitted = job.get("applied_date") or job.get("date_found", "")

        # Build location string
        city = job.get("city", "")
        country = job.get("location", "UK")
        location = f"{city}, {country}" if city else country

        # Data matching Google Sheets columns
        row_data = [
            job.get("company", "Unknown"),           # A: Company Name
            application_status,                       # B: Application Status
            job.get("title", ""),                    # C: Role
            job.get("salary", ""),                   # D: Salary
            date_submitted,                          # E: Date Submitted
            job.get("url", ""),                      # F: Link to Job Req
            "N/A",                                   # G: Rejection Reason (default)
            location,                                # H: Location
            job.get("deadline", ""),                 # I: Deadline
            job.get("notes", ""),                    # J: Notes
            job.get("ai_summary", ""),               # K: AI Summary
        ]

        # Thin border style
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        # Add data to row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border

        # Make URL clickable (column F = column 6)
        url = job.get("url", "")
        if url:
            url_cell = ws.cell(row=row_num, column=6)
            url_cell.hyperlink = url
            url_cell.value = "View Job"  # Display text instead of full URL
            url_cell.font = Font(color="0563C1", underline="single")  # Blue underlined link

        # Apply alternating row colors (light green / white)
        if row_num % 2 == 0:
            row_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            for col_num in range(1, len(row_data) + 1):
                ws.cell(row=row_num, column=col_num).fill = row_fill
            # Keep link blue even on green rows
            if url:
                ws.cell(row=row_num, column=6).font = Font(color="0563C1", underline="single")

        # Color-code status cell (column B)
        self._color_status_cell(ws, row_num, application_status)

        # Color-code rejection reason cell (column G)
        self._color_rejection_cell(ws, row_num, "N/A")
    
    def _color_status_cell(self, ws, row_num, status):
        """Apply color to status cell based on status (Google Sheets style)"""

        # Colors matching Google Sheets screenshot
        status_colors = {
            "Submitted - Pending Response": "4CAF50",  # Green
            "Have Not Applied": "64B5F6",              # Blue
            "Interview Scheduled": "81C784",           # Light green
            "Offer Received": "2E7D32",                # Dark green
            "Rejected": "E57373",                      # Red
            "N/A": "9E9E9E",                           # Grey
        }

        color = status_colors.get(status, "9E9E9E")
        cell = ws.cell(row=row_num, column=2)  # Status column B
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")  # White text
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _color_rejection_cell(self, ws, row_num, reason):
        """Apply color to rejection reason cell (Google Sheets style)"""

        # Grey background for N/A, red for actual rejections
        if reason == "N/A" or not reason:
            color = "9E9E9E"  # Grey
        else:
            color = "E57373"  # Red

        cell = ws.cell(row=row_num, column=7)  # Rejection Reason column G
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _update_status_cell(self, ws, row_num, status):
        """Update status cell with new status and color"""

        # Map internal status to display status
        status_map = {
            "new": "Have Not Applied",
            "liked": "Have Not Applied",
            "maybe": "Have Not Applied",
            "disliked": "Have Not Applied",
            "applied": "Submitted - Pending Response",
            "interview": "Interview Scheduled",
            "offer": "Offer Received",
            "rejected": "Rejected"
        }

        display_status = status_map.get(status, "Have Not Applied")
        cell = ws.cell(row=row_num, column=2)
        cell.value = display_status

        # Update color
        self._color_status_cell(ws, row_num, display_status)
    
    def _add_fancy_dropdowns(self, ws):
        """Add dropdowns matching Google Sheets format"""

        max_row = max(ws.max_row, 100)  # Extend dropdowns for future entries

        # Application Status dropdown (column B)
        status_options = '"Submitted - Pending Response,Have Not Applied,Interview Scheduled,Offer Received,Rejected,N/A"'
        status_dv = DataValidation(type="list", formula1=status_options, allow_blank=False)
        status_dv.error = "Please select a valid status"
        status_dv.errorTitle = "Invalid Status"
        status_dv.prompt = "Select application status"
        status_dv.promptTitle = "Application Status"
        ws.add_data_validation(status_dv)
        status_dv.add(f"B2:B{max_row}")

        # Rejection Reason dropdown (column G)
        rejection_options = '"N/A,No Response,Position Filled,Not Qualified,Failed Interview,Salary Mismatch,Location Issue,Other"'
        rejection_dv = DataValidation(type="list", formula1=rejection_options, allow_blank=True)
        rejection_dv.error = "Please select a valid reason"
        rejection_dv.errorTitle = "Invalid Reason"
        rejection_dv.prompt = "Select rejection reason (if applicable)"
        rejection_dv.promptTitle = "Rejection Reason"
        ws.add_data_validation(rejection_dv)
        rejection_dv.add(f"G2:G{max_row}")
    
    def _auto_fit_columns(self, ws):
        """Auto-fit column widths matching Google Sheets layout"""

        column_widths = {
            "A": 25,   # Company Name
            "B": 28,   # Application Status
            "C": 45,   # Role
            "D": 18,   # Salary
            "E": 14,   # Date Submitted
            "F": 40,   # Link to Job Req
            "G": 18,   # Rejection Reason
            "H": 20,   # Location
            "I": 14,   # Deadline
            "J": 30,   # Notes
            "K": 50,   # AI Summary
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width


# CLI Interface
if __name__ == "__main__":
    import sys
    
    tracker = EnhancedJobTracker()
    
    if len(sys.argv) > 1:
        command = sys.argv[1]
        
        if command == "export":
            tracker.export_to_excel_fancy()
        
        elif command == "stats":
            print("\n📊 Job Tracker Stats:\n")
            print(f"   Total jobs: {len(tracker.jobs)}")
            print(f"   👍 Liked: {len(tracker.get_jobs_by_status('liked'))}")
            print(f"   🤔 Maybe: {len(tracker.get_jobs_by_status('maybe'))}")
            print(f"   👎 Disliked: {len(tracker.get_jobs_by_status('disliked'))}")
            print(f"   📤 Applied: {len(tracker.get_jobs_by_status('applied'))}")
            print(f"   🎤 Interview: {len(tracker.get_jobs_by_status('interview'))}")
            print(f"   🎉 Offer: {len(tracker.get_jobs_by_status('offer'))}")
            print()
        
        elif command == "add":
            # Manual entry
            print("\n➕ Add Manual Job Entry\n")
            
            title = input("Job Title: ")
            company = input("Company: ")
            city = input("City: ")
            location = input("Country [UK]: ") or "UK"
            url = input("URL: ")
            job_type = input("Type (Industry/PhD) [Industry]: ") or "Industry"
            salary = input("Salary [Not specified]: ") or "Not specified"
            deadline = input("Deadline [Not specified]: ") or "Not specified"
            notes = input("Notes: ")
            
            job = {
                "title": title,
                "company": company,
                "location": location,
                "city": city,
                "url": url,
                "type": job_type,
                "job_type": job_type,
                "salary": salary,
                "deadline": deadline,
                "notes": notes,
                "requirements": [],
                "expectations": [],
                "post_date": "Today",
                "cv_required": "Not specified",
                "cover_letter_required": "Not specified",
            }
            
            tracker.add_job(job, status="new")
            tracker.export_to_excel_fancy()
            
            print("\n✅ Job added and spreadsheet updated!\n")
    
    else:
        print("\nUsage:")
        print("  python tracker_enhanced.py export  - Export to Excel")
        print("  python tracker_enhanced.py stats   - View statistics")
        print("  python tracker_enhanced.py add     - Add manual entry")
        print()